"""PDF and Excel export for NR UE Capability reports."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

_OUTPUT = Path(__file__).resolve().parent.parent / "data" / "reports"


def _safe_stem(name: str) -> str:
    stem = Path(name).stem[:36]
    return re.sub(r"[^\w\-]", "_", stem) or "uecap"


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _step_status_label(step: dict[str, Any]) -> str:
    st = step.get("status")
    if st == "found":
        return "PASS"
    if st == "optional":
        return "N/A"
    return "FAIL"


def export_ue_capability_excel(report: dict[str, Any]) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed. pip install openpyxl"}

    _OUTPUT.mkdir(parents=True, exist_ok=True)
    ts = _timestamp()
    stem = _safe_stem(report.get("filename") or "uecap")
    filename = f"nr_ue_capability_{stem}_{ts}.xlsx"
    out = _OUTPUT / filename

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    green = PatternFill("solid", fgColor="DCFCE7")
    red = PatternFill("solid", fgColor="FEE2E2")
    amber = PatternFill("solid", fgColor="FEF3C7")
    blue = PatternFill("solid", fgColor="E0E7FF")

    ws_sum["A1"] = "NR UE Capability Report"
    ws_sum["A1"].font = header_font
    ws_sum["A2"] = f"Source log: {report.get('filename', '—')}"
    ws_sum["A3"] = f"Generated (UTC): {datetime.now(timezone.utc).isoformat(timespec='seconds')}"

    rows = [
        ("Overall", report.get("overall")),
        ("Procedure steps", f"{report.get('procedure_passed')}/{report.get('procedure_total')}"),
        ("Segmentation", "Yes" if report.get("segmentation") else "No"),
        ("Bands detected", ", ".join(report.get("bands_detected") or []) or "—"),
    ]
    fm = report.get("first_missing")
    if fm:
        rows.append(("First gap", fm.get("label")))
        rows.append(("Note", fm.get("note")))

    r = 5
    for label, value in rows:
        ws_sum.cell(row=r, column=1, value=label).font = label_font
        ws_sum.cell(row=r, column=2, value=value or "")
        r += 1

    overall = report.get("overall")
    status_cell = ws_sum.cell(row=5, column=2)
    if overall == "COMPLETE":
        status_cell.fill = green
    elif overall == "PARTIAL":
        status_cell.fill = amber
    elif overall == "FIELDS_ONLY":
        status_cell.fill = blue
    elif overall:
        status_cell.fill = red

    r += 1
    ws_sum.cell(row=r, column=1, value="Alerts").font = label_font
    r += 1
    for alert in report.get("alerts") or []:
        ws_sum.cell(row=r, column=1, value=alert)
        ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 60

    ws_proc = wb.create_sheet("Procedure")
    proc_headers = ["Step", "Direction", "Status", "Note"]
    for col, h in enumerate(proc_headers, start=1):
        ws_proc.cell(row=1, column=col, value=h).font = label_font

    for i, step in enumerate(report.get("steps") or [], start=2):
        status = _step_status_label(step)
        ws_proc.cell(row=i, column=1, value=step.get("label"))
        ws_proc.cell(row=i, column=2, value=step.get("direction"))
        sc = ws_proc.cell(row=i, column=3, value=status)
        if status == "PASS":
            sc.fill = green
        elif status == "FAIL":
            sc.fill = red
        ws_proc.cell(row=i, column=4, value=step.get("note") or "")

    for col, width in zip("ABCD", [36, 14, 10, 48]):
        ws_proc.column_dimensions[col].width = width

    ws_ie = wb.create_sheet("IE hints")
    ws_ie.cell(row=1, column=1, value="Field").font = label_font
    ws_ie.cell(row=1, column=2, value="Status").font = label_font
    for i, field in enumerate(report.get("fields") or [], start=2):
        ws_ie.cell(row=i, column=1, value=field.get("field"))
        st = "FOUND" if field.get("status") == "found" else "NOT SEEN"
        ws_ie.cell(row=i, column=2, value=st)
    ws_ie.column_dimensions["A"].width = 24
    ws_ie.column_dimensions["B"].width = 14

    ts_rows = report.get("troubleshooting") or []
    if ts_rows:
        ws_ts = wb.create_sheet("Troubleshooting")
        ws_ts.cell(row=1, column=1, value="Symptom").font = label_font
        ws_ts.cell(row=1, column=2, value="Checks").font = label_font
        row = 2
        for item in ts_rows:
            ws_ts.cell(row=row, column=1, value=item.get("symptom"))
            checks = item.get("checks") or []
            ws_ts.cell(row=row, column=2, value="\n".join(f"• {c}" for c in checks))
            row += 1
        ws_ts.column_dimensions["A"].width = 40
        ws_ts.column_dimensions["B"].width = 70

    refs = report.get("references") or []
    specs = report.get("spec_refs") or []
    if refs or specs:
        ws_ref = wb.create_sheet("References")
        ws_ref.cell(row=1, column=1, value="Type").font = label_font
        ws_ref.cell(row=1, column=2, value="Value").font = label_font
        row = 2
        for url in refs:
            ws_ref.cell(row=row, column=1, value="URL")
            ws_ref.cell(row=row, column=2, value=url)
            row += 1
        for spec in specs:
            ws_ref.cell(row=row, column=1, value="3GPP")
            ws_ref.cell(row=row, column=2, value=spec)
            row += 1
        ws_ref.column_dimensions["B"].width = 80

    wb.save(out)
    return {
        "ok": True,
        "filename": filename,
        "download_url": f"/api/reports/{filename}",
        "type": "excel",
    }


def export_ue_capability_pdf(report: dict[str, Any]) -> dict[str, Any]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return {"ok": False, "error": "reportlab not installed. pip install reportlab"}

    _OUTPUT.mkdir(parents=True, exist_ok=True)
    ts = _timestamp()
    stem = _safe_stem(report.get("filename") or "uecap")
    filename = f"nr_ue_capability_{stem}_{ts}.pdf"
    out = _OUTPUT / filename

    doc = SimpleDocTemplate(
        str(out),
        pagesize=letter,
        rightMargin=0.6 * inch,
        leftMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("UeCapTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=8)
    body = styles["Normal"]
    small = ParagraphStyle("UeCapSmall", parent=body, fontSize=9, leading=11)

    story: list[Any] = []
    story.append(Paragraph("NR UE Capability Report", title_style))
    story.append(
        Paragraph(
            f"<b>Log:</b> {report.get('filename', '—')} &nbsp; "
            f"<b>Result:</b> {report.get('overall', '—')} "
            f"({report.get('procedure_passed')}/{report.get('procedure_total')} steps)",
            body,
        )
    )
    if report.get("segmentation"):
        story.append(Paragraph("<b>RRC segmentation:</b> detected", small))
    bands = report.get("bands_detected") or []
    if bands:
        story.append(Paragraph(f"<b>Bands in log:</b> {', '.join(bands)}", small))
    story.append(
        Paragraph(
            f"Generated (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}",
            small,
        )
    )
    story.append(Spacer(1, 10))

    fm = report.get("first_missing")
    if fm and report.get("overall") != "COMPLETE":
        story.append(Paragraph(f"<b>First gap:</b> {fm.get('label', '')}", body))
        if fm.get("note"):
            story.append(Paragraph(fm["note"], small))
        story.append(Spacer(1, 8))

    for alert in report.get("alerts") or []:
        story.append(Paragraph(f"• {alert}", small))
    if report.get("alerts"):
        story.append(Spacer(1, 8))

    proc_data = [["Step", "Direction", "Status"]]
    for step in report.get("steps") or []:
        proc_data.append([
            step.get("label") or "",
            step.get("direction") or "",
            _step_status_label(step),
        ])

    proc_tbl = Table(proc_data, colWidths=[3.2 * inch, 1.3 * inch, 0.7 * inch], repeatRows=1)
    proc_tbl.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ])
    )
    story.append(proc_tbl)
    story.append(Spacer(1, 10))

    fields_found = [f.get("field") for f in (report.get("fields") or []) if f.get("status") == "found"]
    if fields_found:
        story.append(Paragraph(f"<b>IE hints:</b> {', '.join(fields_found)}", small))
        story.append(Spacer(1, 8))

    for item in report.get("troubleshooting") or []:
        story.append(Paragraph(f"<b>{item.get('symptom', '')}</b>", small))
        for check in item.get("checks") or []:
            story.append(Paragraph(f"  - {check}", small))
        story.append(Spacer(1, 4))

    refs = report.get("references") or []
    specs = report.get("spec_refs") or []
    if refs or specs:
        story.append(Spacer(1, 6))
        story.append(Paragraph("<b>References</b>", body))
        for url in refs:
            story.append(Paragraph(url, small))
        if specs:
            story.append(Paragraph(f"Specs: {', '.join(specs)}", small))

    doc.build(story)
    return {
        "ok": True,
        "filename": filename,
        "download_url": f"/api/reports/{filename}",
        "type": "pdf",
    }


def export_ue_capability_reports(report: dict[str, Any]) -> dict[str, Any]:
    if not report.get("ok", True):
        return {"xlsx": {"ok": False, "error": "Invalid report"}, "pdf": {"ok": False, "error": "Invalid report"}}

    return {
        "xlsx": export_ue_capability_excel(report),
        "pdf": export_ue_capability_pdf(report),
    }
