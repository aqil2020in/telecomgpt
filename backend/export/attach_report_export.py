"""PDF and Excel export for NR SA Initial Attach reports."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

_OUTPUT = Path(__file__).resolve().parent.parent / "data" / "reports"


def _safe_stem(name: str) -> str:
    stem = Path(name).stem[:36]
    return re.sub(r"[^\w\-]", "_", stem) or "attach"


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def export_attach_excel(report: dict[str, Any]) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed. pip install openpyxl"}

    _OUTPUT.mkdir(parents=True, exist_ok=True)
    ts = _timestamp()
    stem = _safe_stem(report.get("filename") or "attach")
    filename = f"nr_sa_attach_{stem}_{ts}.xlsx"
    out = _OUTPUT / filename

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    green = PatternFill("solid", fgColor="DCFCE7")
    red = PatternFill("solid", fgColor="FEE2E2")
    amber = PatternFill("solid", fgColor="FEF3C7")

    ws_sum["A1"] = "NR SA Initial Attach Report"
    ws_sum["A1"].font = header_font
    ws_sum["A2"] = f"Source log: {report.get('filename', '—')}"
    ws_sum["A3"] = f"Generated (UTC): {datetime.now(timezone.utc).isoformat(timespec='seconds')}"

    rows = [
        ("Overall", report.get("overall")),
        ("Steps passed", f"{report.get('passed')}/{report.get('total')}"),
    ]
    fm = report.get("first_missing")
    if fm:
        rows.append(("First gap", fm.get("label")))
        rows.append(("Troubleshooting", fm.get("fail_hint")))

    r = 5
    for label, value in rows:
        ws_sum.cell(row=r, column=1, value=label).font = label_font
        ws_sum.cell(row=r, column=2, value=value or "")
        r += 1

    overall = report.get("overall")
    status_cell = ws_sum.cell(row=5, column=2)
    if overall == "COMPLETE":
        status_cell.fill = green
    elif overall == "PARTIAL":
        status_cell.fill = amber
    elif overall:
        status_cell.fill = red

    r += 1
    ws_sum.cell(row=r, column=1, value="Alerts").font = label_font
    r += 1
    alerts = report.get("alerts") or []
    if alerts:
        for alert in alerts:
            ws_sum.cell(row=r, column=1, value=alert)
            ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1
    else:
        ws_sum.cell(row=r, column=1, value="None")
        r += 1

    log_summary = report.get("log_summary") or {}
    r += 1
    ws_sum.cell(row=r, column=1, value="Log stats").font = label_font
    r += 1
    ws_sum.cell(row=r, column=1, value=f"Lines: {log_summary.get('total_lines', '—')}")
    r += 1
    ws_sum.cell(row=r, column=1, value=f"Errors: {log_summary.get('error_count', '—')}")

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 60

    ws_chk = wb.create_sheet("Checklist")
    headers = ["Phase", "Step", "Status", "Matched pattern", "Fail hint"]
    for col, h in enumerate(headers, start=1):
        cell = ws_chk.cell(row=1, column=col, value=h)
        cell.font = label_font

    for i, step in enumerate(report.get("steps") or [], start=2):
        status = "PASS" if step.get("status") == "found" else "FAIL"
        ws_chk.cell(row=i, column=1, value=step.get("phase"))
        ws_chk.cell(row=i, column=2, value=step.get("label"))
        sc = ws_chk.cell(row=i, column=3, value=status)
        sc.fill = green if status == "PASS" else red
        ws_chk.cell(row=i, column=4, value=step.get("matched") or "")
        ws_chk.cell(row=i, column=5, value=step.get("fail_hint") or "")

    for col, width in zip("ABCDE", [12, 42, 10, 28, 48]):
        ws_chk.column_dimensions[col].width = width

    refs = report.get("references") or []
    if refs:
        ws_ref = wb.create_sheet("References")
        ws_ref.cell(row=1, column=1, value="URL").font = label_font
        for i, url in enumerate(refs, start=2):
            ws_ref.cell(row=i, column=1, value=url)
        ws_ref.column_dimensions["A"].width = 80

    top_errors = (log_summary.get("top_errors") or []) if log_summary else []
    if top_errors:
        ws_err = wb.create_sheet("Top errors")
        ws_err.cell(row=1, column=1, value="Count").font = label_font
        ws_err.cell(row=1, column=2, value="Message").font = label_font
        for i, err in enumerate(top_errors, start=2):
            ws_err.cell(row=i, column=1, value=err.get("count"))
            ws_err.cell(row=i, column=2, value=err.get("message"))
        ws_err.column_dimensions["B"].width = 90

    wb.save(out)
    return {
        "ok": True,
        "filename": filename,
        "download_url": f"/api/reports/{filename}",
        "type": "excel",
    }


def export_attach_pdf(report: dict[str, Any]) -> dict[str, Any]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return {"ok": False, "error": "reportlab not installed. pip install reportlab"}

    _OUTPUT.mkdir(parents=True, exist_ok=True)
    ts = _timestamp()
    stem = _safe_stem(report.get("filename") or "attach")
    filename = f"nr_sa_attach_{stem}_{ts}.pdf"
    out = _OUTPUT / filename

    doc = SimpleDocTemplate(
        str(out),
        pagesize=letter,
        rightMargin=0.6 * inch,
        leftMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AttachTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=8,
    )
    body = styles["Normal"]
    small = ParagraphStyle("Small", parent=body, fontSize=9, leading=11)

    story: list[Any] = []
    story.append(Paragraph("NR SA Initial Attach Report", title_style))
    story.append(
        Paragraph(
            f"<b>Log:</b> {report.get('filename', '—')} &nbsp; "
            f"<b>Result:</b> {report.get('overall', '—')} "
            f"({report.get('passed')}/{report.get('total')} steps)",
            body,
        )
    )
    story.append(
        Paragraph(
            f"Generated (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}",
            small,
        )
    )
    story.append(Spacer(1, 10))

    fm = report.get("first_missing")
    if fm and report.get("overall") != "COMPLETE":
        story.append(Paragraph(f"<b>First gap:</b> {fm.get('label', '')}", body))
        if fm.get("fail_hint"):
            story.append(Paragraph(fm["fail_hint"], small))
        story.append(Spacer(1, 8))

    alerts = report.get("alerts") or []
    if alerts:
        story.append(Paragraph("<b>Alerts</b>", body))
        for alert in alerts:
            story.append(Paragraph(f"• {alert}", small))
        story.append(Spacer(1, 8))

    table_data = [["Phase", "Step", "Status"]]
    for step in report.get("steps") or []:
        status = "PASS" if step.get("status") == "found" else "FAIL"
        table_data.append([step.get("phase") or "", step.get("label") or "", status])

    tbl = Table(table_data, colWidths=[1.1 * inch, 4.2 * inch, 0.7 * inch], repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    for i, step in enumerate(report.get("steps") or [], start=1):
        if step.get("status") == "found":
            tbl.setStyle(TableStyle([("BACKGROUND", (2, i), (2, i), colors.HexColor("#dcfce7"))]))
        else:
            tbl.setStyle(TableStyle([("BACKGROUND", (2, i), (2, i), colors.HexColor("#fee2e2"))]))

    story.append(tbl)
    story.append(Spacer(1, 12))

    log_summary = report.get("log_summary") or {}
    if log_summary:
        story.append(
            Paragraph(
                f"<b>Log stats:</b> {log_summary.get('total_lines', '—')} lines, "
                f"{log_summary.get('error_count', '—')} errors",
                small,
            )
        )
        story.append(Spacer(1, 8))

    refs = report.get("references") or []
    if refs:
        story.append(Paragraph("<b>References</b>", body))
        for url in refs:
            story.append(Paragraph(url, small))

    doc.build(story)
    return {
        "ok": True,
        "filename": filename,
        "download_url": f"/api/reports/{filename}",
        "type": "pdf",
    }


def export_attach_reports(report: dict[str, Any]) -> dict[str, Any]:
    """Generate Excel and PDF; skip gracefully if a dependency is missing."""
    if not report.get("ok", True):
        return {"xlsx": {"ok": False, "error": "Invalid report"}, "pdf": {"ok": False, "error": "Invalid report"}}

    xlsx = export_attach_excel(report)
    pdf = export_attach_pdf(report)
    return {"xlsx": xlsx, "pdf": pdf}
